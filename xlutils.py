'''
Script: xlutils.py
Author: G. Paxton
Purpose: To aid in reading/writing excel files
Revision: March 2018
'''

import sys
import os
import logging

import openpyxl

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border,
from openpyxl.styles import Side, Alignment, Protection, Font
from openpyxl.workbook.defined_name import DefinedName,DefinedNameList
from openpyxl.worksheet.table import Table, TableStyleInfo

# Data
'''
Fonts
'''
defaultFont = Font(name='Calibri', size=11, bold=False,
                   italic=False, vertAlign=None, underline='none',
                   strike=False, color='FF000000')
'''
Fills
'''
defaultFill = PatternFill(fill_type=None, start_color='FFFFFFFF',
                          end_color='FF000000')

'''
Alignment
'''
defaultAlign=Alignment(horizontal='general', vertical='bottom',
                 text_rotation=0, wrap_text=False,
                 shrink_to_fit=False, indent=0)
centerAlign = Alignment(horizontal="center", vertical="center")

'''
Format
'''
defaultFormat = 'General'
commaFormat = '#,##0.00'

'''
Borders
'''
defaultBorder = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                diagonal=Side(border_style=None,
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                               color='FF000000')
               )

# CLASS
class XLUtil:
    '''Class for reading and writing an xlsx file
    '''

    def __init__(self, excelPath, logFile='XLUtil_log.txt'):
        '''Initialize XL with a path
        '''
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.xlFile = self.xlDir = ""
        
        # set up logging
        logging.basicConfig(
            filename=logFile, filemode='w', level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s')
        
        # get the directory path and file name
        head, tail = os.path.split(excelPath)
        
        # check if directory exists or if wasn't given
        if(os.path.isdir(head) == True or head == ""):
            # check for extenstion .xlsx
            if(excelPath[-5:] != ".xlsx"):
                tail+='.xlsx'
            # check if file exists or needs created
            if(os.path.isfile(tail) == True):
                logging.debug('Load workbook ' + tail)        
	        self.workbook = openpyxl.load_workbook(tail)
                self.worksheet = self.workbook.active
            else:
                logging.debug('Starting workbook ' + tail)
            self.xlDir = head
            self.xlFile = tail
        else:
            print('\n{0}- Directory does not exist\n'.format(head))

            
    def save_workbook(self, fileName=self.xlFile):
        '''Save the excel workbook

        Give the fileName to save under
        default is xlFile name picked from excel path
        '''
        print("Saving {0}".format(fileName))
        self.workbook.save(fileName)


    def make_sheet(self, sheetName, index=0):
        """Create a new worksheet

        @param sheetName: name of new worksheet
        @param index: position of new worksheet (default 0)
        """
        self.workbook.create_sheet(sheetName, index)


    def get_sheets(self):
        """Return a list of all the worksheets
        """
        return self.workbook.sheetnames

    
    def select_sheet(self, sheetName):
        """Set the sheetName as the active worksheet
        """
        if sheetName not in get_sheets():
            make_sheet(sheetname)
            
        self.worksheet = self.workbook[sheetName]

        
    def remove_sheet(self, sheetName):
        """Remove the sheet from the workbook
        """
        if sheetName in get_sheets():
            self.workbook.remove(self.workbook[sheetName])

            
    def write(self, column, row, value):
        """Write to a cell in the active sheet

        @param column : column value as a number (col A = 1)
        @param row : row value as a number
        @param value : data to write to cell
        """
        self.worksheet.cell(row=row, column=column).value = value


'''
UNDER CONSTRUCTION


    def write_row(self, column, row, values):
        """Write to the cells in a row

        @param column : column value as a number to start (A = 1)
        @param 


    ##########
    # write a list of values along a single row
    ##########
    def write_row(self, colStart, row, values):
        for i in range(0, len(values)):
            self.write_cell(colStart+i, row, values[i])

    ##########
    # write a list of values along a single column
    ##########
    def write_col(self, col, rowStart, values):
        for i in range(0, len(values)):
            self.write_cell(col, rowStart+i, values[i])

    ##########
    # read the value of a cell
    ##########
    def read_cell(self, col, row):
        return self.worksheet.cell(row=row, column=col).value

    ##########
    # read a list of values from a row of cells
    ##########
    def read_row(self, colStart, row, length):
        data = []
        for i in range(0, length):
            data.append(self.read_cell(colStart+i, row))
        return data

    ##########
    # read a list of values from a column of cells
    ##########
    def read_col(self, col, rowStart, length):
        data = []
        for i in range(0, length):
            data.append(self.read_cell(col, rowStart+i))
        return data

    ##########
    # freeze the top and side panels, specifying their sizes by the corner cell
    ##########
    def freeze(self, col, row):
        col_letter = get_column_letter(col)
        coord = col_letter + str(row)
#        print(coord)
        self.worksheet.freeze_panes = coord

    ##########
    # format a cell by applying font, alignment, number style, and color filling
    ##########
    def format_cell(self, col, row,
                    fnt=dfFont, algn=dfAlgn,
                    num=dfNum, fll=dfFill):
        cell = self.worksheet.cell(row=row, column=col)
        cell.font = fnt
        cell.alignment = algn
        cell.number_format = num
        cell.fill = fll

    ##########
    # format an entire row of cells
    ##########
    def format_row(self, col, row, length,
                   fnt=dfFont, algn=dfAlgn,
                   num=dfNum, fll=dfFill):
        for i in range(0, length):
            self.format_cell(col+i, row, fnt, algn, num, fll)

    ##########
    # format a column of cells
    ##########
    def format_col(self, col, row, length,
                   fnt=dfFont, algn=dfAlgn,
                   num=dfNum, fll=dfFill):
        for i in range(0, length):
            self.format_cell(col, row+i, fnt, algn, num, fll)

    ##########
    # set the width of a column, if auto is made true the column will auto size to the largest cell value
    ##########
    def set_col_width(self, col, w=10, auto=False):
        col_letter = get_column_letter(col)
        logging.debug(col_letter)
        if(auto ==True):
            max_length = 0
            column = self.worksheet[col_letter]
            for cell in column:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if(max_length > 50):
                max_length = 50
            adjusted_width = (max_length + 2) * 1.2
            w = adjusted_width            
        self.worksheet.column_dimensions[col_letter].width = w       

    ##########
    # return a list of names that were defined for a region of cells
    ##########
    def get_ranges(self):
#        print(self.workbook.defined_names.localnames(self.worksheet))
        return self.workbook.defined_names.definedName

    ##########
    # check if a given region name exists in the defined name's list
    ##########
    def is_table_exists(self, tableName):
        # get the definedName's list, then get all the name properties of each definedName
        dfns = self.get_ranges()
        dfnNames = []
        for dfn in dfns:
            dfnNames.append(dfn.name)
            logging.debug(dfn.name)
            logging.debug(dfn.attr_text)
        #Check if the given name exists
        if(tableName not in dfnNames):
            logging.debug("Table Not Found - " + tableName)
            return False
        else:
            logging.debug("Table Found - " + tableName)
            return True

    ##########
    # Make a attr_text value for a defined name from the active sheet and the starting and ending coordinates of the region
    ##########
    def build_attr_text(self, coords):
        # active sheet
        text = "$"+self.worksheet.title+"."
        # starting Col and row
        text += "$"+coords[0][0] +"$" + coords[0][1] + ":"
        # ending Col and row
        text += "$"+coords[1][0] + "$" + coords[1][1]
        logging.debug(text)
        return text

    ##########
    # From the defined name attr_text value
    # Make a list of the sheet the region exists, starting col, starting row, ending col, ending row 
    ##########
    def debuild_attr_text(self, attr_text):
        logging.debug(attr_text)
        leave = attr_text[1:] # remove first $
        findList = ['$', '$', ':', '$','$']
        texts = []
        for char in findList:
            idx = leave.find(char)
            remove = leave[:idx]
            leave = leave[idx+1:]
            logging.debug(char +'\t'+remove+'\t'+leave)
            texts.append(remove)
        texts.append(leave)
        texts[0] = texts[0][:-1]
        del texts[3]
        logging.debug(texts)
        return texts

    ##########
    # Create a sortable table and label the region by using the defined name properties
    # if the table already exists, rewrite all values and update the size
    ##########
    def write_table(self, col, row, tableData, tableName):
        # table size properties
        numRows = len(tableData)
        numCols = len(tableData[0])
        startCoord = [get_column_letter(col),str(row)]
        endCoord=[get_column_letter(col+numCols-1),
                  str(row+numRows-1)]
        span =startCoord[0]+startCoord[1]+':'+endCoord[0]+endCoord[1]
        coords = [startCoord, endCoord]
        # write the table
        for i in range(0, numRows):
            self.write_row(col, row+i, tableData[i])
        # make filterable
#        self.worksheet.auto_filter.ref = span

        # Check if table exists, if not create one
        if(self.is_table_exists(tableName) == False):
            dfn = DefinedName(name=tableName)
            dfn.attr_text = self.build_attr_text(coords)
            self.workbook.defined_names.append(dfn)
        else:
            foundTable = self.workbook.defined_names[tableName]
            foundTable.attr_text = self.build_attr_text(coords)

    ##########
    # Add a row of data to an existing table by looking up the defined name,
    # inserting the row at the end, then updating the region size
    ##########
    def append_table_row(self, rowData, tableName):
        if(self.is_table_exists(tableName) == True):
            foundTable = self.workbook.defined_names[tableName]
            rawCoords=self.debuild_attr_text(foundTable.attr_text)  
            sheet = rawCoords[0]
            startCol = rawCoords[1]
            startRow = rawCoords[2]
            endCol = rawCoords[3]
            endRow = rawCoords[4]
            col = column_index_from_string(startCol)
            row = int(endRow)+1
            self.write_row(col, row, rowData)
            startCoords = [startCol, startRow]
            endCoords = [endCol, str(row)]
            coords = [startCoords, endCoords]
            foundTable.attr_text=self.build_attr_text(coords)

    ##########
    # look up the table by the defined name and format the header, then the rows
    ##########
    def format_table(self, tableName):
        if(self.is_table_exists(tableName) == True):
            foundTable = self.workbook.defined_names[tableName]
            rawCoords=self.debuild_attr_text(foundTable.attr_text)  
            sheet = rawCoords[0]
            startCol = column_index_from_string(rawCoords[1])
            startRow = int(rawCoords[2])
            endCol = column_index_from_string(rawCoords[3])
            endRow = int(rawCoords[4])
            numCols = endCol - startCol + 1
            numRows = endRow - startRow + 1
            # remove formatting
            for i in range(0, numRows):
                self.format_row(startCol, startRow, numCols)
            # format header
            self.format_row(startCol, startRow, numCols,
                            fnt=boldFont, algn=centerAlign)
            # format rows
            for i in range(0, numRows-1):
                if(i%2 == 0):
                    fill = greyFill
                else:
                    fill = whiteFill
                self.format_row(startCol, startRow+1+i, numCols,
                                algn=centerAlign, fll=fill)
