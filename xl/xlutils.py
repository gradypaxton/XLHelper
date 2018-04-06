'''
Script: xlutils.py
Author: G. Paxton
Purpose: To aid in reading/writing excel files
Revision: March 2018
'''
###############################################################################
import sys
import os
import logging
import re

import openpyxl

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border
from openpyxl.styles import Side, Alignment, Protection, Font
from openpyxl.workbook.defined_name import DefinedName,DefinedNameList
from openpyxl.worksheet.table import Table, TableStyleInfo

# Data
'''
Fonts
'''
FONT = Font(name='Calibri', size=11, bold=False, italic=False,
            vertAlign=None, underline='none', strike=False,
            color='FF000000')
FONT_BOLD = Font(bold=True)
'''
Fills
'''
FILL = PatternFill(fill_type=None, start_color='FFFFFFFF',
                   end_color='FF000000')
FILL_GREY = PatternFill('solid', fgColor='DDDDDD')
FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')
'''
Alignment
'''
ALIGN =Alignment(horizontal='general', vertical='bottom',
                 text_rotation=0, wrap_text=False,
                 shrink_to_fit=False, indent=0)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

'''
Format
'''
FORMAT = 'General'
FORMAT_COMMA = '#,##0.00'

'''
Borders
'''
BORDER = Border(left=Side(border_style=None, color='FF000000'),
                right=Side(border_style=None,color='FF000000'),
                top=Side(border_style=None,color='FF000000'),
                bottom=Side(border_style=None, color='FF000000'),
                diagonal=Side(border_style=None, color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None, color='FF000000'),
                vertical=Side(border_style=None, color='FF000000'),
                horizontal=Side(border_style=None, color='FF000000')
               )


# CLASS
class XLUtil:
    '''Class for reading and writing an xlsx file

    This class is particularily good for saving time by incorporating
    the active worksheet into each function so it doesn't have to 
    keep being specified  
    '''

    def __init__(self, excelPath, logFile='XLUtil_log.txt'):
        '''Initialize XL with a path

        '''
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.xlFile = self.xlDir = ""
        
        # set up logging
        logging.basicConfig(
            filename=logFile, filemode='w', level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s')
        
        # get the directory path and file name
        head, tail = os.path.split(excelPath)
        
        # check if directory exists or if wasn't given
        if(os.path.isdir(head) == True or head == ""):
            # check for extenstion .xlsx
            if(excelPath[-5:] != ".xlsx"):
                tail+='.xlsx'
            # check if file exists or needs created
            if(os.path.isfile(tail) == True):
                logging.debug('Load workbook ' + tail)        
	        self.workbook = openpyxl.load_workbook(tail)
                self.worksheet = self.workbook.active
            else:
                logging.debug('Starting workbook ' + tail)
            self.xlDir = head
            self.xlFile = tail
        else:
            print('\n{0}- Directory does not exist\n'.format(head))

            
    def save_workbook(self, fileName=None):
        '''Save the excel workbook

        Give the fileName to save under
        default is xlFile name picked from excel path
        '''
        if fileName is None:
            fileName=self.xlFile
        print("Saving {0}".format(fileName))
        self.workbook.save(fileName)


    # WORKSHEET METHODS
    def get_sheets(self):
        """Return a list of all the worksheets

        """
        return self.workbook.sheetnames

    
    def get_active_sheet(self):
        """Return the active worksheet

        """
        return self.worksheet

    
    def make_sheet(self, sheetName, index=0):
        """Create a new worksheet

        @param sheetName: name of new worksheet
        @param index: position of new worksheet (default 0)
        """
        self.workbook.create_sheet(sheetName, index)

    
    def select_sheet(self, sheetName):
        """Set the sheetName as the active worksheet

        @param sheetName: name of the worksheet to work in
        """
        if sheetName not in self.get_sheets():
            make_sheet(sheetname)
            
        self.worksheet = self.workbook[sheetName]

        
    def remove_sheet(self, sheetName):
        """Remove the sheet from the workbook

        @param sheetName: name of the worksheet to remove
        """
        if sheetName in self.get_sheets():
            self.workbook.remove(self.workbook[sheetName])

    def copy_sheet(self, src, dest):
        """Copy a sheet

        @param src : sheet to be copied
        @param dst : sheet to copy to
        """
        dest = this.workbook.copy_worksheet(src)

        
    def rename_sheet(self, newName):
        """Rename the active worksheet

        @param newName : string
        """
        self.worksheet.title = newName

        
    # CELL VALUES
    def get_coord(self, coordinate):
        """Parse the coordinate string for column & row value

        @param coordinate : string value of coordinate (eg "B2")
        @return [column, row] : numeric values

        Note: Does not catch no match (search returns NoneType)
        """
        # match any alpha char [a-z], one or more times (+) at start ^
        col = re.search('^[a-zA-Z]+', coordinate).group()
        
        # match any digit (\d), one or more times (+) at end $
        row = re.search('\d+$', coordinate).group()
        
        return [column_index_from_string(col), int(row)]

    
    def make_coord(self, column, row):
        """Make a coordinate string from a row & column integer

        @param column : integer value of column
        @param row : integer value of row
        @return coord : string of format 'A1"
        """                
        return get_column_letter(column)+str(row)
                

    def get_span(self, coords):
        """Parse a span of coordinates for row & column values

        @param coords : string of coords (eg "A1:B2")
        @return [col1, row1], [col2, row2] : ints of cols & rows
        """
        idx = coords.index(':')
        coord1 = self.get_coord(coords[0:idx])
        coord2 = self.get_coord(coords[idx+1:])
        return coord1, coord2

    
    def make_span(self, coord1, coord2):
        """Make a span of coordinates from row & column values

        @param coord1 : list [column, row] of starting numeric vals
        @param coord2 : list [column, row] of ending numeric vals
        @return span : string of form "A1:B2"
        """
        coord1 = self.make_coord(coord1[0], coord1[1])
        coord2 = self.make_coord(coord2[0], coord2[1])
        return coord1 + ':' + coord2

    
    # READING AND WRITING METHODS
    def write(self, coord, value):
        """Write to a cell in the active sheet

        @param coord : string of format 'A1"
        @param value : data to write to cell
        """
        column, row = self.get_coord(coord)
        self.worksheet.cell(row=row, column=column).value = value

        
    def read(self, coord):
        """Read from a cell in the active sheet

        @param coord : string of format 'A1"
        """
        column, row = self.get_coord(coord)
        return self.worksheet.cell(row=row, column=column).value

    
    def write_row(self, coord, values):
        """Write to the cells in a row

        @param coord : string of format 'A1"
        @param values : a list of values to write
        """
        column, row = self.get_coord(coord)
        for i in range(len(values)):
            self.write(self.make_coord(column+i, row), values[i])

            
    def read_row(self, coord, length):
        """Read the cells of a row

        @param coord : string of format 'A1"
        @param length : the number of cells to read
        @return values : list of values read
        """
        column, row = self.get_coord(coord)
        return [self.read( self.make_coord(column+i, row) )
                for i in range(length)]

    def write_column(self, coord, values):
        """Write to the cells in a column

        @param coord : string of format 'A1"
        @param values : a list of values to write
        """
        column, row = self.get_coord(coord)
        for i in range(len(values)):            
            self.write(self.make_coord(column, row+i), values[i])
    
            
    def read_column(self, coord, length):
        """Read the cells of a row

        @param coord : string of format 'A1"
        @param length : the number of cells to read
        @return values : list of values read
        """
        column, row = self.get_coord(coord)
        return [self.read( self.make_coord(column, row+i) )
                for i in range(length)]

    
    def write_block(self, coord, values):
        """Write to the cells in both columns and rows

        @param coord : starting cell, string of format "A1"
        @param values : list of lists to write, [ [1,2], [3,4] ]

        Note: values don't have to be equal in length
        """
        column, row = self.get_coord(coord)
        for i in range(len(values)):
            for j in range(len(values[i])):
                self.write(self.make_coord(column+j, row+i),
                           values[i][j])

                
    def read_block(self, span):
        """Read from the cells specified by the span

        @param span : string of format 'A1:B2'
        @return table : list of lists, list of rows read
        """
        ([col1, row1], [col2, row2]) = self.get_span(span)        
        table_width = col2 - col1 + 1
        table_height = row2 - row1 + 1
        return [self.read_row(self.make_coord(col1, row1+i),
                              table_width)
                for i in range(table_height)]

    
    def append_row(self, values):
        """Append a row to the last row of the active sheet

        @param values : list of values
        """
        self.worksheet.append(row)

        
    #FORMATING METHODS
    def merge(self, span):
        """Merge a span of cells to create one cell

        @param span : string value of cells ('A1:B2')
        """
        self.worksheet.merge_cells(span)


    def unmerge(self, span):
        """Unmerge a span of cells to create many cells

        @param span : string value of cells ('A1:B2')
        """
        self.worksheet.unmerge_cells(span)

        
    def style(self, coord, font=FONT, align=ALIGN,
               num=FORMAT, fill=FILL):
        """Style a cell with font, alignment, fill, and format

        @param coord : string value of cell ('A1')
        @param span : string value of cells ('A1:B2')
        @param font : the font class of openpyxl
        @param align : the align class of openpyxl
        @param num : the format class of openpyxl
        @param fill : the fill class of openpyxl
        """
        column, row = self.get_coord(coord)
        cell = self.worksheet.cell(row=row, column=column)
        cell.font = font
        cell.alignment = align
        cell.number_format = num
        cell.fill = fill


    def style_block(self, span, font=FONT, align=ALIGN,
                    num=FORMAT, fill=FILL):
        """Style a span of cells

        @param span : string value of cells ('A1:B2')
        @param font : the font class of openpyxl
        @param align : the align class of openpyxl
        @param num : the format class of openpyxl
        @param fill : the fill class of openpyxl
        """
        ([col1, row1], [col2, row2]) = self.get_span(span)
        for i in range(col2-col1+1):
            for j in range(row2 - row1 + 1):
                self.style(self.make_coord(col1 + i, row1 + j),
                           font, align, num, fill)

            
    def freeze(self, coord):
        """Freeze the rows and columns before the cell value given

        @param coord : string value of cell ('A1')
        """
        self.worksheet.freeze_panes = coord


    def set_column_width(self, column, w=10, auto=False):
        """Set the width of a column or autosize it (but not > 50)

        @param column : column to size
        @param width : width (in ?) to set
        @param auto : auto size column True/False
        """
        column_letter = get_column_letter(column)
        if(auto == True):
            max_length = 0
            column = self.worksheet[column_letter]
            for cell in column:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if(max_length > 50):
                max_length = 50
            adjusted_width = (max_length + 2) * 1.1
            w = adjusted_width            
        self.worksheet.column_dimensions[column_letter].width = w 

'''
UNDER CONSTRUCTION

    def get_ranges(self):
#        print(self.workbook.defined_names.localnames(self.worksheet))
        return self.workbook.defined_names.definedName

    ##########
    # check if a given region name exists in the defined name's list
    ##########
    def is_table_exists(self, tableName):
        # get the definedName's list, then get all the name properties of each definedName
        dfns = self.get_ranges()
        dfnNames = []
        for dfn in dfns:
            dfnNames.append(dfn.name)
            logging.debug(dfn.name)
            logging.debug(dfn.attr_text)
        #Check if the given name exists
        if(tableName not in dfnNames):
            logging.debug("Table Not Found - " + tableName)
            return False
        else:
            logging.debug("Table Found - " + tableName)
            return True

    ##########
    # Make a attr_text value for a defined name from the active sheet and the starting and ending coordinates of the region
    ##########
    def build_attr_text(self, coords):
        # active sheet
        text = "$"+self.worksheet.title+"."
        # starting Column and row
        text += "$"+coords[0][0] +"$" + coords[0][1] + ":"
        # ending Column and row
        text += "$"+coords[1][0] + "$" + coords[1][1]
        logging.debug(text)
        return text

    ##########
    # From the defined name attr_text value
    # Make a list of the sheet the region exists, starting column, starting row, ending column, ending row 
    ##########
    def debuild_attr_text(self, attr_text):
        logging.debug(attr_text)
        leave = attr_text[1:] # remove first $
        findList = ['$', '$', ':', '$','$']
        texts = []
        for char in findList:
            idx = leave.find(char)
            remove = leave[:idx]
            leave = leave[idx+1:]
            logging.debug(char +'\t'+remove+'\t'+leave)
            texts.append(remove)
        texts.append(leave)
        texts[0] = texts[0][:-1]
        del texts[3]
        logging.debug(texts)
        return texts

    ##########
    # Create a sortable table and label the region by using the defined name properties
    # if the table already exists, rewrite all values and update the size
    ##########
    def write_table(self, column, row, tableData, tableName):
        # table size properties
        numRows = len(tableData)
        numColumns = len(tableData[0])
        startCoord = [get_columnumn_letter(column),str(row)]
        endCoord=[get_columnumn_letter(column+numColumns-1),
                  str(row+numRows-1)]
        span =startCoord[0]+startCoord[1]+':'+endCoord[0]+endCoord[1]
        coords = [startCoord, endCoord]
        # write the table
        for i in range(0, numRows):
            self.write_row(column, row+i, tableData[i])
        # make filterable
#        self.worksheet.auto_filter.ref = span

        # Check if table exists, if not create one
        if(self.is_table_exists(tableName) == False):
            dfn = DefinedName(name=tableName)
            dfn.attr_text = self.build_attr_text(coords)
            self.workbook.defined_names.append(dfn)
        else:
            foundTable = self.workbook.defined_names[tableName]
            foundTable.attr_text = self.build_attr_text(coords)

    ##########
    # Add a row of data to an existing table by looking up the defined name,
    # inserting the row at the end, then updating the region size
    ##########
    def append_table_row(self, rowData, tableName):
        if(self.is_table_exists(tableName) == True):
            foundTable = self.workbook.defined_names[tableName]
            rawCoords=self.debuild_attr_text(foundTable.attr_text)  
            sheet = rawCoords[0]
            columnumn = rawCoords[1]
            row = rawCoords[2]
            endColumn = rawCoords[3]
            endRow = rawCoords[4]
            column = columnumn_index_from_string(columnumn)
            row = int(endRow)+1
            self.write_row(column, row, rowData)
            startCoords = [columnumn, row]
            endCoords = [endColumn, str(row)]
            coords = [startCoords, endCoords]
            foundTable.attr_text=self.build_attr_text(coords)

    ##########
    # look up the table by the defined name and format the header, then the rows
    ##########
    def format_table(self, tableName):
        if(self.is_table_exists(tableName) == True):
            foundTable = self.workbook.defined_names[tableName]
            rawCoords=self.debuild_attr_text(foundTable.attr_text)  
            sheet = rawCoords[0]
            columnumn = columnumn_index_from_string(rawCoords[1])
            row = int(rawCoords[2])
            endColumn = columnumn_index_from_string(rawCoords[3])
            endRow = int(rawCoords[4])
            numColumns = endColumn - columnumn + 1
            numRows = endRow - row + 1
            # remove formatting
            for i in range(0, numRows):
                self.format_row(columnumn, row, numColumns)
            # format header
            self.format_row(columnumn, row, numColumns,
                            fnt=boldFont, algn=centerAlign)
            # format rows
            for i in range(0, numRows-1):
                if(i%2 == 0):
                    fill = greyFill
                else:
                    fill = whiteFill
                self.format_row(columnumn, row+1+i, numColumns,
                                algn=centerAlign, fll=fill)


TODO

"""Charts"""
This is a definite branch worthy feature
'''
